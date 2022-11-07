#! /usr/bin/env python3
# -*- coding:utf-8 -*-
# author@zhaopeng
# @time: 2022/10/27 23:41

from openpyxl import load_workbook

# https://zhuanlan.zhihu.com/p/409954748
# https://geek-docs.com/python/python-tutorial/python-openpyxl.html
'''
读取excel信息
'''
wb = load_workbook('./demo.xlsx')
# 获取sheet列表
print(wb.sheetnames)

# 加载sheet
ws = wb["表三"]
# 读取单个单元格
print(ws['A1'].value)
# 读取多个单元格
calls = ws['B4': 'N21']
for i, j in calls:
    print(i.value, j.value)
